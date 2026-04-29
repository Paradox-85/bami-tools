#!/usr/bin/env python3
"""
equipment_unpivot.py — Unpivot Equipment Characteristics Excel
Usage:
    python equipment_unpivot.py <input_file> [--mapping <file>] [--output <file>]
"""

import argparse, sys, os
from pathlib import Path
import yaml

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

def load_yaml_config(config_path="config.yaml"):
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)

def find_header_row(ws, max_scan=5):
    for i, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True)):
        if row and row[0] and str(row[0]).strip().upper() == "TAG CODE":
            return i
    return None


def build_column_map(header_row):
    columns, skip = [], set()
    for i, cell in enumerate(header_row):
        if cell is None or i in skip:
            continue
        hdr = str(cell).strip()
        if hdr.upper() == "TAG CODE":
            continue
        uom_idx = None
        if i + 1 < len(header_row):
            nxt = header_row[i + 1]
            if nxt and str(nxt).strip().upper() in ("UOM", "U.O.M"):
                uom_idx = i + 1
                skip.add(i + 1)
        columns.append({"attr_name": hdr, "col_idx": i, "uom_col_idx": uom_idx})
    return columns


def load_mapping(mapping_file):
    if not mapping_file or not os.path.exists(mapping_file):
        return {}
    mapping, path = {}, Path(mapping_file)
    if path.suffix.lower() in (".xlsx", ".xls"):
        wb = load_workbook(mapping_file, read_only=True, data_only=True)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                mapping[str(row[0]).strip()] = str(row[1]).strip()
    elif path.suffix.lower() == ".csv":
        import csv
        with open(mapping_file, newline="", encoding="utf-8-sig") as f:
            for row in list(csv.reader(f))[1:]:
                if len(row) >= 2 and row[0]:
                    mapping[row[0].strip()] = row[1].strip()
    return mapping


def process_sheet(ws, mapping=None):
    mapping = mapping or {}
    hdr_idx = find_header_row(ws)
    if hdr_idx is None:
        print(f"  WARN: No TAG CODE in '{ws.title}', skipping.")
        return []
    all_rows = list(ws.iter_rows(values_only=True))
    col_map = build_column_map(all_rows[hdr_idx])
    records = []
    for row in all_rows[hdr_idx + 1:]:
        if not row or row[0] is None:
            continue
        tag = str(row[0]).strip()
        if not tag:
            continue
        for col in col_map:
            val = row[col["col_idx"]] if col["col_idx"] < len(row) else None
            uom = row[col["uom_col_idx"]] if col["uom_col_idx"] and col["uom_col_idx"] < len(row) else None
            val_str = str(val).strip() if val is not None else ""
            if not val_str:
                continue
            records.append({
                "Tag Name": tag,
                "Attribute Name": mapping.get(col["attr_name"], col["attr_name"]),
                "Attribute Value": val_str,
                "Attribute UoM": str(uom).strip() if uom else "",
            })
    return records


def main():
    parser = argparse.ArgumentParser(description="Unpivot Equipment Characteristics Excel")
    parser.add_argument("input_file", nargs="?", default=None)
    parser.add_argument("--config", "-c", default=None, help="Path to config.yaml")
    parser.add_argument("--output", "-o", default=None)
    parser.add_argument("--mapping", "-m", default=None)
    parser.add_argument("--sheet-prefix", default="5")
    args = parser.parse_args()

    cfg = {}
    config_path = args.config or "config.yaml"
    if os.path.exists(config_path):
        cfg = load_yaml_config(config_path)
        print(f"Config loaded: {config_path}")

    input_file   = args.input_file   or cfg.get("input_file")
    output_file  = args.output       or cfg.get("output_file")
    mapping_file = args.mapping      or cfg.get("mapping_file")
    prefix       = args.sheet_prefix or cfg.get("sheet_prefix", "5")  # ← вычислено

    if not input_file:
        print("ERROR: input_file not specified (CLI arg or config.yaml)")
        sys.exit(1)

    inp = Path(input_file)
    out = Path(output_file) if output_file else inp.with_name(inp.stem + "_unpivot.xlsx")

    if not inp.exists():
        print(f"ERROR: Input file not found: {inp}")
        sys.exit(1)

    mapping = load_mapping(mapping_file)
    print(f"Input : {inp}\nOutput: {out}")

    wb = load_workbook(inp, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s.startswith(prefix)]  # ← prefix, не args.sheet_prefix
    print(f"Sheets ({len(sheets)}): {sheets}")

    all_records = []
    for name in sheets:
        recs = process_sheet(wb[name], mapping)
        for r in recs:
            r["Source Sheet"] = name
        all_records.extend(recs)
        print(f"  {name}: {len(recs)} rows")
    wb.close()

    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "Unpivot Result"
    headers = ["Tag Name", "Attribute Name", "Attribute Value", "Attribute UoM", "Source Sheet"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E5F8A")
        cell.alignment = Alignment(horizontal="center")
    for rec in all_records:
        ws.append([rec["Tag Name"], rec["Attribute Name"],
                   rec["Attribute Value"], rec["Attribute UoM"], rec["Source Sheet"]])
    for col, w in zip("ABCDE", [20, 40, 30, 15, 30]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    out_wb.save(out)
    print(f"\n✓ Done! {len(all_records)} records → {out}")


if __name__ == "__main__":
    main()